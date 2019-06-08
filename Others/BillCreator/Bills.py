from tkinter import *
from tkinter import messagebox
import os
import openpyxl
import inflect
import re
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border,Side,BORDER_THIN
from openpyxl.utils import column_index_from_string

p=inflect.engine()

row = 9
col = 2

wb = openpyxl.load_workbook("my.xlsx")
sheet = wb['Sheet1']
thin_border = Border(left= Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

def calc():
    global row
    summ = 0
    for i in range(9,row-7):
        summ += sheet.cell(row = i, column = 6).value
    gst = round(summ*9/100)
    total = round(summ + 2*gst)
    add_value(summ,row-6,6)
    add_value(gst,row-5,6)
    add_value(gst,row-4,6)
    add_value(total,row-2,6)
    add_value((p.number_to_words(total)).title(),row-1,1)

def add_border():
    global row
    for r in range(7,row+5):
       for c in range(1,7):
           sheet.cell(row = r, column = c).border = thin_border
           

def add_value(value,row,col):
    cell = sheet.cell(row = row, column = col)
    cell.value = value
    cell.alignment = Alignment(wrap_text=True)
    font_style(cell,'Times New Roman',True)

def font_style(cell,fontname,b,f_size = 11):
    cell.font = Font(name = fontname,bold = b, size = f_size )

def align(cell,h_align,v_align):
        cell.alignment = Alignment(wrap_text=True,horizontal=h_align, vertical=v_align )



def style():
    global row, t2,t3,t4,t5

    date = t2.get()
    noc = t3.get()
    cgstno = t4.get()
    cpan = t5.get()
    sno = t6.get()

    add_value(date,2,6)
    add_value(noc,3,3)
    add_value('Clients GST No.:- ' + str(cgstno),row+1,5)
    add_value('Pan No.:- ' + str(cpan),row+2,5)
    add_value(sno,5,3)

    sheet.merge_cells('A' + str(row-2) + ':C' + str(row-2))
    cell1 = sheet.cell(row = row-2, column = 1)
    align(cell1,'right','center')
    font_style(cell1,'Times New Roman',True)
    
    add_value(8,row,6)

    sheet.merge_cells('A' + str(row-1) + ':F' + str(row-1))
    cell1 = sheet.cell(row = row-1, column = 1)
    align(cell1,'center','center')
    cell1.font = Font(name = 'Times New Roman',bold = True, size = 11, color='ffffff' )
    cell1.fill = PatternFill(bgColor='FFF000',fill_type='solid')

    sheet.merge_cells('A' + str(row) + ':F' + str(row))
    cell1 = sheet.cell(row = row, column = 1)
    align(cell1,'center','center')
    font_style(cell1,'Times New Roman',True)

    for i in range(1,8):
        cell1 = sheet.cell(row = row+1, column = i)
        cell2 = sheet.cell(row = row+2, column = i)
        font_style(cell1,'Times New Roman',True)
        font_style(cell2,'Times New Roman',True)
    
    sheet.merge_cells('B' + str(row+1) + ':C' + str(row+1))
    sheet.merge_cells('E' + str(row+1) + ':F' + str(row+1))
    sheet.merge_cells('B' + str(row+2) + ':C' + str(row+2))
    sheet.merge_cells('E' + str(row+2) + ':F' + str(row+2))

    
    sheet.merge_cells('A' + str(row+3) + ':F' + str(row+4))
    cell1 = sheet.cell(row = row+3, column = 1)
    font_style(cell1,'Times New Roman',False,7)
    cell1.alignment = Alignment(wrap_text=True)
    align(cell1,'center','center')

    add_border()





def copy():
    copy_sheet = wb['Sheet2']
    global row
    for r in copy_sheet.iter_rows():
        for c in r:

            x = sheet.cell(row = row + 5, column = c.column) 
            x.value = c.value
            
        row += 1
    calc()
    style()

no_rows = 0
# 14.75 25.5

def add_row():
    def next_row():
        try:
            global row,col
            sheet.insert_rows(row)
            sheet.merge_cells('B' + str(row) + ':C' + str(row))
            row_title = e.get()
            quantity = (e1.get())
            q = float(re.findall('[0-9]*[.]?[0-9]*',quantity)[0])
            rate = (e2.get())
            r = float(re.findall('[0-9]*[.]?[0-9]*',rate)[0])
            sheet.row_dimensions[row].height = (len(row_title)//42 + 1)*14.75
            add_value(row_title,row,col)
            add_value(quantity,row,col+2)
            add_value(rate,row,col+3)
            add_value(q*r,row,col+4)
            
            row += 1
            b2.destroy()
            b3.destroy()
            e1.destroy()
            e2.destroy()
            e.destroy()
            l1 = Label(root,text = row_title)
            l2 = Label(root,text = quantity)
            l3 = Label(root,text = rate)
            l1.grid(row = no_rows+14, column = 6)
            l2.grid(row = no_rows+14, column = 10)
            l3.grid(row = no_rows+14, column = 14)
            add_row()
        except ValueError:
            messagebox.showwarning("Unsuccessfull","Enter an integer value!!")
    
    def done():


        name = t1.get()
        date = t2.get()
        noc = t3.get()
        cgstno = t4.get()
        cpan = t5.get()
        sno = t6.get()
        if name == '' or date == '' or noc == '' or cgstno == '' or cpan == '' or sno == '' :
            messagebox.showwarning("Unsuccessfull","Enter the above fields properly!!!")
        elif os.path.isfile(name):
            messagebox.showwarning("Unsuccessfull","File already exists")
        
        else:
            global row
            copy()
            wb.remove(wb['Sheet2'])
            if not os.path.exists('C:\Bills'):
                os.makedirs('C:\Bills')
            wb.save('C:\Bills\\' + name + '.xlsx')
            root.destroy()
    

    global no_rows
    no_rows += 1 
    b1.destroy()
    ln1 = Label(root, text = 'Description')
    ln2 = Label(root, text = 'Quantity')
    ln3 = Label(root, text = 'Rate')

    l = Label(root,text = str(no_rows))
    e = Entry(root)
    e1 = Entry(root)
    e2 = Entry(root)
    b2 = Button(root, text = "Save and Next", command = next_row)
    b3 = Button(root, text = "Done", command = done)
    
    ln1.grid(row = 13, column = 6)
    ln2.grid(row = 13, column = 10)
    ln3.grid(row = 13, column = 14)

    l.grid(row = no_rows+14, column = 2)
    e.grid(row = no_rows+14, column = 6)
    e1.grid(row = no_rows+14, column = 10)
    e2.grid(row = no_rows+14, column = 14)
    b2.grid(row = no_rows+14 , column = 18)
    b3.grid(row = no_rows+15 , column = 10)




root = Tk()

root.title("Bill")
root.geometry('500x400')
l1 = Label(root, text = 'Filename')
l2 = Label(root, text = 'Date')
l3 = Label(root, text = 'Name Of Client')
l4 = Label(root, text = 'Client GST No')
l5 = Label(root, text = 'Client Pan No')
l6 = Label(root, text = 'Serial Number Of This Bill')

t1 = Entry(root)
t2 = Entry(root)
t3 = Entry(root)
t4 = Entry(root)
t5 = Entry(root)
t6 = Entry(root)

b1 = Button(root, text = 'Add A Row', command = add_row)

l1.grid(row = 2, column = 6)
l2.grid(row = 4, column = 6)
l3.grid(row = 6, column = 6)
l4.grid(row = 8, column = 6)
l5.grid(row = 10, column = 6)
l6.grid(row = 12, column = 6)

t1.grid(row = 2, column = 14)
t2.grid(row = 4, column = 14)
t3.grid(row = 6, column = 14)
t4.grid(row = 8, column = 14)
t5.grid(row = 10, column = 14)
t6.grid(row = 12, column = 14)

b1.grid(row = 13, column = 5)
root.mainloop()